"""lg-isin — resident LangServer for global listed-company data ingest.

Tools:
  listed.ingest.usSecurities        — SEC EDGAR ticker list + OpenFIGI → vertex_isin_security
  listed.ingest.jpSecurities        — OpenFIGI JP ticker range (TSE) → vertex_isin_security
  listed.enrich.cik                 — EDGAR CIK submissions → enrich exchange_mic/sic
  listed.ingest.edinetFiling        — EDINET filing list → vertex_isin_filing
  listed.normalize.linkLeiIsin      — cross-reference vertex_isin_security ↔ vertex_open_lei_entity
  listed.normalize.linkSecurityFiling — link vertex_isin_security → vertex_isin_filing by ticker
  listed.news.fetchPressReleases    — RSS/PR feed fetch for a company
  listed.ingest.irPage              — fetch company IR HTML page → vertex_isin_ir_doc
  listed.ingest.irPdf               — download + extract PDF → vertex_isin_ir_doc
  listed.ingest.irExcel             — download + extract Excel → vertex_isin_ir_doc
  listed.ingest.irWord              — download + extract Word → vertex_isin_ir_doc
  listed.coverage.tick              — coverage count snapshot

Cron endpoints:
  POST /cron/bootstrap        — one-shot: full US sweep then full JP sweep (background)
  POST /cron/sweep-us         — daily rolling cursor (500/run)
  POST /cron/sweep-jp         — daily rolling cursor (50/run)
  POST /cron/enrich-cik       — weekly CIK enrichment
  POST /cron/link-lei         — weekly LEI-ISIN linkage
  POST /cron/link-filing      — weekly security-filing edge linkage
  POST /cron/coverage         — coverage tick

Table schema:
  vertex_isin_security / vertex_isin_filing / vertex_isin_ir_doc
  edge_isin_lei_match / edge_isin_security_filing
  mv_isin_security_country_cnt / mv_isin_filing_ticker_cnt / mv_isin_ir_doc_type_cnt
Apply migrations:
  20260515120000_alter_vertex_isin_tables.up.sql
  20260515130000_vertex_isin_ir_doc.up.sql
  20260515140000_isin_graph_wiring.up.sql
"""

from __future__ import annotations

import asyncio
import hashlib
import io
import os
import time
import uuid
from typing import Any

import aiohttp
import asyncpg
from fastapi import FastAPI, HTTPException
import uvicorn


_OWNER_DID = "did:web:isin.etzhayyim.com"
_EDGAR_UA = "isin.etzhayyim.com/1.0 contact@etzhayyim"
_EDGAR_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
_OPENFIGI_URL = "https://api.openfigi.com/v3/mapping"
_OPENFIGI_FILTER_URL = "https://api.openfigi.com/v3/filter"
_EDINET_URL = "https://api.edinet-fsa.go.jp/api/v2/documents.json"
_EDINET_DOC_URL = "https://api.edinet-fsa.go.jp/api/v2/documents/{doc_id}"
_ASX_LIST_URL = "https://www.asx.com.au/asx/research/ASXListedCompanies.csv"

_EXCH_TO_COUNTRY: dict[str, str] = {
    "GY": "DE", "FP": "FR", "NA": "NL", "BB": "BE", "IM": "IT",
    "LN": "GB", "SW": "CH", "HK": "HK", "AU": "AU",
    "KS": "KR", "KQ": "KR", "SP": "SG", "TT": "TW",
}
_EU_EXCH_CODES = ["GY", "FP", "NA", "BB", "IM", "LN"]

DB_URL = os.environ.get(
    "DATABASE_URL",
    "REDACTED_USE_DATABASE_URL_ENV",
)
EDINET_KEY = os.environ.get("EDINET_SUBSCRIPTION_KEY", "")

_EXCH_TO_MIC: dict[str, str] = {
    "NYSE": "XNYS",
    "NASDAQ": "XNAS",
    "NYSE MKT": "XASE",
    "NYSE Arca": "ARCX",
    "CBOE": "XCBO",
    "OTC": "OTCM",
}
_EDINET_FORM_NAMES: dict[str, str] = {
    "030000": "有価証券報告書",
    "043000": "四半期報告書",
    "020000": "臨時報告書",
    "050000": "半期報告書",
}
_FORM_SLUG: dict[str, str] = {
    "有価証券報告書": "yukasho",
    "四半期報告書": "shihankiho",
    "臨時報告書": "rinjiho",
    "半期報告書": "hankiho",
}


def _utc_now() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def _sec_vid(rkey: str) -> str:
    return f"at://{_OWNER_DID}/com.etzhayyim.apps.isin.security/{rkey}"


def _filing_vid(rkey: str) -> str:
    return f"at://{_OWNER_DID}/com.etzhayyim.apps.isin.filing/{rkey}"


def _is_dup(e: Exception) -> bool:
    return isinstance(e, asyncpg.UniqueViolationError)


async def _db() -> asyncpg.Connection:
    return await asyncpg.connect(DB_URL)


# ---------------------------------------------------------------------------
# listed.ingest.usSecurities
# ---------------------------------------------------------------------------

async def tool_ingest_us_securities(
    offset: int = 0,
    limit: int = 200,
    enrichFigi: bool = True,
) -> dict:
    off = max(0, int(offset or 0))
    lim = max(1, min(int(limit or 200), 500))

    async with aiohttp.ClientSession() as session:
        async with session.get(
            _EDGAR_TICKERS_URL,
            headers={"User-Agent": _EDGAR_UA},
            timeout=aiohttp.ClientTimeout(total=30),
        ) as resp:
            if resp.status != 200:
                return {"error": f"EDGAR tickers {resp.status}", "registered": 0}
            raw = await resp.json(content_type=None)

    all_tickers = list(raw.values())
    batch = all_tickers[off: off + lim]
    if not batch:
        return {"ok": True, "registered": 0, "total": len(all_tickers), "exhausted": True}

    figi_map: dict[str, dict] = {}
    if enrichFigi:
        FIGI_BATCH = 100
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(batch), FIGI_BATCH):
                chunk = [
                    {"idType": "TICKER", "idValue": t["ticker"], "exchCode": "US", "marketSecDes": "Equity"}
                    for t in batch[i: i + FIGI_BATCH]
                ]
                try:
                    async with session.post(
                        _OPENFIGI_URL,
                        json=chunk,
                        headers={"Content-Type": "application/json"},
                        timeout=aiohttp.ClientTimeout(total=20),
                    ) as r:
                        if r.status == 200:
                            results = await r.json()
                            for j, entry in enumerate(results):
                                match = (entry.get("data") or [{}])[0]
                                if match and (i + j) < len(batch):
                                    figi_map[batch[i + j]["ticker"]] = match
                except Exception:
                    pass
                if i + FIGI_BATCH < len(batch):
                    await asyncio.sleep(2.6)

    now = _utc_now()
    registered = skipped = errors = 0
    conn = await _db()
    try:
        for edgar in batch:
            cik = str(edgar.get("cik_str", ""))
            ticker = str(edgar.get("ticker", ""))
            name = str(edgar.get("title", ""))
            figi = figi_map.get(ticker, {})
            rkey = f"us-{cik}"
            vid = _sec_vid(rkey)
            asset_class = (
                "equity"
                if figi.get("marketSector") == "Equity"
                else (figi.get("marketSector") or "equity").lower()
            )
            try:
                await conn.execute(
                    "INSERT INTO vertex_isin_security "
                    "(vertex_id, rkey, isin, figi, composite_figi, ticker, cik, name, "
                    "country_code, asset_class, security_type, exch_code, isin_status, "
                    "status, source_did, actor_did, org_did, collected_at, created_at) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18,$19)",
                    vid, rkey, "",
                    figi.get("figi", ""), figi.get("compositeFIGI", ""),
                    ticker, cik, name,
                    "US", asset_class,
                    figi.get("securityType", "Common Stock"),
                    figi.get("exchCode", "US"),
                    "pending", "active",
                    "did:web:isin.etzhayyim.com:source:sec",
                    _OWNER_DID, "anon", now, now,
                )
                registered += 1
            except Exception as e:
                if _is_dup(e):
                    skipped += 1
                else:
                    errors += 1
    finally:
        await conn.close()

    return {
        "ok": True,
        "registered": registered,
        "skipped": skipped,
        "errors": errors,
        "figiEnriched": len(figi_map),
        "total": len(all_tickers),
        "offset": off,
        "limit": lim,
        "nextOffset": off + len(batch),
        "exhausted": len(batch) < lim,
    }


# ---------------------------------------------------------------------------
# listed.ingest.jpSecurities
# ---------------------------------------------------------------------------

async def tool_ingest_jp_securities(
    fromTicker: int = 1000,
    count: int = 25,
) -> dict:
    start = max(1000, min(int(fromTicker or 1000), 9999))
    cnt = max(1, min(int(count or 25), 50))
    tickers = [str(t) for t in range(start, min(start + cnt, 10000))]
    if not tickers:
        return {"ok": True, "registered": 0, "exhausted": True}

    FIGI_BATCH = 10
    now = _utc_now()
    registered = skipped = errors = 0

    conn = await _db()
    try:
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(tickers), FIGI_BATCH):
                chunk_tickers = tickers[i: i + FIGI_BATCH]
                payload = [
                    {"idType": "TICKER", "idValue": t, "exchCode": "JP", "marketSecDes": "Equity"}
                    for t in chunk_tickers
                ]
                try:
                    async with session.post(
                        _OPENFIGI_URL,
                        json=payload,
                        headers={"Content-Type": "application/json"},
                        timeout=aiohttp.ClientTimeout(total=20),
                    ) as resp:
                        if resp.status != 200:
                            continue
                        results = await resp.json()
                except Exception:
                    continue

                for j, entry in enumerate(results):
                    match = (entry.get("data") or [{}])[0]
                    if not match or not match.get("name"):
                        continue
                    ticker = chunk_tickers[j]
                    rkey = f"jp-{ticker}"
                    vid = _sec_vid(rkey)
                    try:
                        await conn.execute(
                            "INSERT INTO vertex_isin_security "
                            "(vertex_id, rkey, isin, figi, composite_figi, ticker, name, "
                            "country_code, asset_class, security_type, exch_code, isin_status, "
                            "status, source_did, actor_did, org_did, collected_at, created_at) "
                            "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)",
                            vid, rkey, "",
                            match.get("figi", ""), match.get("compositeFIGI", ""),
                            ticker, match.get("name", ""),
                            "JP", "equity",
                            match.get("securityType", "Common Stock"),
                            "JP", "pending", "active",
                            "did:web:isin.etzhayyim.com:source:openfigi",
                            _OWNER_DID, "anon", now, now,
                        )
                        registered += 1
                    except Exception as e:
                        if _is_dup(e):
                            skipped += 1
                        else:
                            errors += 1

                if i + FIGI_BATCH < len(tickers):
                    await asyncio.sleep(2.6)
    finally:
        await conn.close()

    next_from = (start + cnt) if (start + cnt) <= 9999 else 1000
    return {
        "ok": True,
        "registered": registered,
        "skipped": skipped,
        "errors": errors,
        "fromTicker": start,
        "count": cnt,
        "nextFrom": next_from,
        "exhausted": (start + cnt) > 9999,
    }


# ---------------------------------------------------------------------------
# listed.enrich.cik
# ---------------------------------------------------------------------------

async def tool_enrich_cik(cik: int = 0) -> dict:
    if not cik:
        return {"error": "cik required"}
    cik_int = int(cik)
    cik_str = str(cik_int).zfill(10)

    async with aiohttp.ClientSession() as session:
        async with session.get(
            f"https://data.sec.gov/submissions/CIK{cik_str}.json",
            headers={"User-Agent": _EDGAR_UA},
            timeout=aiohttp.ClientTimeout(total=20),
        ) as resp:
            if resp.status != 200:
                return {"error": f"EDGAR CIK {resp.status}", "cik": cik_int}
            facts = await resp.json(content_type=None)

    exchanges: list[str] = facts.get("exchanges") or []
    sic = str(facts.get("sic") or "")
    sic_desc = str(facts.get("sicDescription") or "")
    mic = _EXCH_TO_MIC.get(exchanges[0] if exchanges else "", "")
    rkey = f"us-{cik_int}"
    now = _utc_now()

    conn = await _db()
    try:
        await conn.execute(
            "UPDATE vertex_isin_security "
            "SET exchange_mic=$1, sic=$2, sic_desc=$3, collected_at=$4 "
            "WHERE rkey=$5",
            mic, sic, sic_desc, now, rkey,
        )
    except Exception as e:
        return {"error": f"DB update failed: {e}", "cik": cik_int}
    finally:
        await conn.close()

    return {
        "ok": True,
        "cik": cik_int,
        "tickers": facts.get("tickers") or [],
        "exchanges": exchanges,
        "mic": mic,
        "sic": sic,
        "sicDescription": sic_desc,
    }


# ---------------------------------------------------------------------------
# listed.ingest.edinetFiling
# ---------------------------------------------------------------------------

async def tool_ingest_edinet_filing(
    ticker: str = "",
    edinetCode: str = "",
    subscriptionKey: str = "",
) -> dict:
    ticker = str(ticker or "").strip()
    edinet_code = str(edinetCode or "").strip()
    if not ticker and not edinet_code:
        return {"error": "ticker or edinetCode required"}

    key = subscriptionKey or EDINET_KEY
    if not key:
        return {"ok": True, "blocked": "EDINET_KEY_MISSING", "ticker": ticker or edinet_code}

    ticker_id = ticker or edinet_code
    params: dict[str, str] = {"type": "2"}
    if edinet_code:
        params["edinetCode"] = edinet_code
    elif ticker:
        params["secCode"] = f"{ticker}0"

    headers: dict[str, str] = {"User-Agent": _EDGAR_UA, "Ocp-Apim-Subscription-Key": key}
    qs = "&".join(f"{k}={v}" for k, v in params.items())

    async with aiohttp.ClientSession() as session:
        async with session.get(
            f"{_EDINET_URL}?{qs}",
            headers=headers,
            timeout=aiohttp.ClientTimeout(total=20),
        ) as resp:
            if resp.status != 200:
                return {"ok": True, "ticker": ticker_id, "filings": [], "note": f"EDINET {resp.status}"}
            data = await resp.json(content_type=None)

    docs = [d for d in (data.get("results") or []) if d.get("edinetCode")]
    if not docs:
        return {"ok": True, "ticker": ticker_id, "filings": [], "note": "no EDINET docs"}

    company = docs[0]
    name = company.get("filerName") or ticker_id
    display_ticker = (company.get("secCode") or "").rstrip("0") or ticker

    sorted_docs = sorted(docs, key=lambda d: d.get("submitDateTime") or "", reverse=True)
    filings_by_type: dict[str, dict] = {}
    for doc in sorted_docs:
        form_name = _EDINET_FORM_NAMES.get(doc.get("formCode") or "")
        if form_name and form_name not in filings_by_type:
            filings_by_type[form_name] = {
                "docID": doc.get("docID", ""),
                "form": form_name,
                "period": doc.get("periodEnd") or doc.get("periodStart") or "",
                "submitted": doc.get("submitDateTime") or "",
            }
        if len(filings_by_type) >= 3:
            break

    if not filings_by_type:
        return {"ok": True, "ticker": ticker_id, "name": name, "filings": []}

    now = _utc_now()
    written = errors = 0
    conn = await _db()
    try:
        for filing in filings_by_type.values():
            period_slug = (filing["period"][:10] if filing["period"] else "").replace("-", "")
            form_slug = _FORM_SLUG.get(filing["form"], filing["form"][:6])
            rkey = f"jp-{ticker_id}-{form_slug}-{period_slug}"
            vid = _filing_vid(rkey)
            source_url = (
                f"https://disclosure.edinet-api.go.jp/e01ew/BLMainController.jsp"
                f"?uji.verb=W1E63011CXP01&TID=W1E63011CXP01&documentId={filing['docID']}"
            )
            try:
                await conn.execute(
                    "INSERT INTO vertex_isin_filing "
                    "(vertex_id, rkey, country_code, ticker, edinet_code, doc_id, "
                    "form_type, period_end, submitted_at, source_url, name, "
                    "actor_did, org_did, created_at) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14)",
                    vid, rkey, "JP",
                    display_ticker or ticker_id, edinet_code,
                    filing["docID"], filing["form"],
                    filing["period"][:10] if filing["period"] else None,
                    filing["submitted"][:19] if filing["submitted"] else None,
                    source_url,
                    f"{name} {filing['form']} 期末: {filing['period'][:10]}",
                    _OWNER_DID, "anon", now,
                )
                written += 1
            except Exception as e:
                if _is_dup(e):
                    written += 1
                else:
                    errors += 1
    finally:
        await conn.close()

    return {
        "ok": True,
        "ticker": ticker_id,
        "name": name,
        "displayTicker": display_ticker,
        "written": written,
        "errors": errors,
        "filings": list(filings_by_type.values()),
    }


# ---------------------------------------------------------------------------
# listed.ingest.edinetPdf
# ---------------------------------------------------------------------------

async def tool_ingest_edinet_pdf(
    docId: str,
    ticker: str = "",
    companyName: str = "",
    subscriptionKey: str = "",
) -> dict:
    """Download and extract a specific EDINET document (type=2 PDF) by docId."""
    doc_id = str(docId or "").strip()
    if not doc_id:
        return {"error": "docId required"}
    key = str(subscriptionKey or EDINET_KEY or "").strip()
    if not key:
        return {"ok": True, "blocked": "EDINET_KEY_MISSING", "docId": doc_id}

    full_url = _EDINET_DOC_URL.format(doc_id=doc_id) + "?type=2"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                full_url,
                headers={"User-Agent": _EDGAR_UA, "Ocp-Apim-Subscription-Key": key},
                timeout=aiohttp.ClientTimeout(total=120),
            ) as resp:
                if resp.status != 200:
                    return {"error": f"EDINET doc {resp.status}", "docId": doc_id}
                ct = resp.headers.get("Content-Type", "")
                if "pdf" not in ct.lower() and "octet" not in ct.lower():
                    return {"error": f"unexpected content-type: {ct}", "docId": doc_id}
                data = await resp.read()
    except Exception as e:
        return {"error": str(e), "docId": doc_id}

    def _parse_pdf(pdf_bytes: bytes) -> tuple[str, int]:
        import pdfplumber
        texts: list[str] = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages[:50]:
                t = page.extract_text() or ""
                if t:
                    texts.append(t)
        return "\n".join(texts), page_count

    try:
        content, page_count = await asyncio.to_thread(_parse_pdf, data)
    except Exception as e:
        return {"error": f"PDF parse: {e}", "docId": doc_id}

    title = f"{companyName or ticker or doc_id} EDINET {doc_id}"
    conn = await _db()
    try:
        saved, err = await _save_ir_doc(
            conn, str(ticker or ""), str(companyName or ""),
            "edinet_pdf", full_url, title, content, len(data), page_count,
        )
    finally:
        await conn.close()

    return {
        "ok": True, "docId": doc_id, "ticker": ticker,
        "pages": page_count, "chars": len(content),
        "saved": saved, "note": err,
    }


# ---------------------------------------------------------------------------
# listed.ingest.edinetSweep
# ---------------------------------------------------------------------------

async def tool_ingest_edinet_sweep(
    limit: int = 10,
    subscriptionKey: str = "",
) -> dict:
    """Batch sweep: pull JP tickers from vertex_isin_security, fetch EDINET filing
    metadata, then download 有価証券報告書 PDF if EDINET_SUBSCRIPTION_KEY is set.
    Advances _edinet_cursor each call; resets to 0 when exhausted."""
    global _edinet_cursor
    lim = max(1, min(int(limit or 10), 50))
    key = str(subscriptionKey or EDINET_KEY or "").strip()

    conn = await _db()
    try:
        rows = await conn.fetch(
            f"SELECT ticker, name FROM vertex_isin_security "
            f"WHERE country_code = 'JP' AND ticker IS NOT NULL AND ticker != '' "
            f"ORDER BY created_at ASC LIMIT {lim} OFFSET {_edinet_cursor}"
        )
        total_jp = int(await conn.fetchval(
            "SELECT COUNT(*) FROM vertex_isin_security WHERE country_code = 'JP'"
        ) or 0)
    finally:
        await conn.close()

    if not rows:
        _edinet_cursor = 0
        return {"ok": True, "processed": 0, "exhausted": True, "cursorReset": True}

    filing_ok = filing_skip = pdf_ok = pdf_skip = 0

    for row in rows:
        ticker = row["ticker"]
        company_name = str(row["name"] or "")

        filing_result = await tool_ingest_edinet_filing(ticker=ticker, subscriptionKey=key)

        if filing_result.get("blocked"):
            filing_skip += 1
        elif filing_result.get("written", 0) > 0:
            filing_ok += 1
            if key:
                filings = filing_result.get("filings") or []
                yukasho = next(
                    (f for f in filings if "有価証券報告書" in f.get("form", "")), None
                )
                if yukasho and yukasho.get("docID"):
                    pdf_result = await tool_ingest_edinet_pdf(
                        docId=yukasho["docID"],
                        ticker=ticker,
                        companyName=company_name,
                        subscriptionKey=key,
                    )
                    if pdf_result.get("saved"):
                        pdf_ok += 1
                    elif not pdf_result.get("blocked"):
                        pdf_skip += 1
        else:
            filing_skip += 1

        await asyncio.sleep(0.5)

    _edinet_cursor = (_edinet_cursor + lim) % max(total_jp, 1)
    exhausted = _edinet_cursor == 0

    return {
        "ok": True,
        "processed": len(rows),
        "filingOk": filing_ok, "filingSkip": filing_skip,
        "pdfOk": pdf_ok, "pdfSkip": pdf_skip,
        "cursor": _edinet_cursor, "totalJp": total_jp,
        "exhausted": exhausted,
        "edinetKeyConfigured": bool(key),
    }


# ---------------------------------------------------------------------------
# listed.normalize.linkLeiIsin
# ---------------------------------------------------------------------------

async def tool_link_lei_isin(limit: int = 100) -> dict:
    """Cross-reference vertex_isin_security ↔ vertex_open_lei_entity by company name."""
    lim = max(1, min(int(limit or 100), 500))
    conn = await _db()
    try:
        securities = await conn.fetch(
            f"SELECT vertex_id, name, ticker FROM vertex_isin_security "
            f"WHERE name IS NOT NULL AND name != '' "
            f"ORDER BY created_at DESC LIMIT {lim}"
        )
        linked = skipped = 0
        now = _utc_now()
        for sec in securities:
            lei_row = await conn.fetchrow(
                "SELECT vertex_id FROM vertex_open_lei_entity "
                "WHERE entity_legal_name ILIKE $1 LIMIT 1",
                f"%{sec['name']}%",
            )
            if not lei_row:
                continue
            edge_id = f"at://{_OWNER_DID}/com.etzhayyim.apps.isin.leiMatch/{sec['ticker'] or uuid.uuid4().hex[:8]}"
            try:
                await conn.execute(
                    "INSERT INTO edge_isin_lei_match "
                    "(edge_id, src_vid, dst_vid, match_score, match_method, "
                    "actor_did, org_did, created_at) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7,$8)",
                    edge_id, sec["vertex_id"], lei_row["vertex_id"],
                    0.7, "name_ilike",
                    _OWNER_DID, "anon", now,
                )
                linked += 1
            except Exception as e:
                if _is_dup(e):
                    skipped += 1
    finally:
        await conn.close()

    return {"ok": True, "linked": linked, "skipped": skipped}


# ---------------------------------------------------------------------------
# listed.news.fetchPressReleases
# ---------------------------------------------------------------------------

_PR_FEEDS: dict[str, str] = {
    "prnewswire": "https://www.prnewswire.com/rss/news-releases-list.rss",
    "businesswire": "https://feed.businesswire.com/rss/home/?rss=G1&_gl=1*1*",
    "globenewswire": "https://www.globenewswire.com/RssFeed/subjectcode/23-Technology",
    "jpx_tse": "https://www.jpx.co.jp/English/markets/listing/co/01.html",
}

async def tool_fetch_press_releases(
    ticker: str = "",
    source: str = "prnewswire",
    limit: int = 10,
) -> dict:
    """Fetch recent press releases/news for a company from RSS sources."""
    ticker = str(ticker or "").strip()
    lim = max(1, min(int(limit or 10), 50))
    feed_url = _PR_FEEDS.get(source, _PR_FEEDS["prnewswire"])

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                feed_url,
                headers={"User-Agent": _EDGAR_UA},
                timeout=aiohttp.ClientTimeout(total=15),
            ) as resp:
                if resp.status != 200:
                    return {"ok": False, "error": f"feed {resp.status}", "source": source}
                text = await resp.text()
    except Exception as e:
        return {"ok": False, "error": str(e), "source": source}

    import re
    items = re.findall(r"<item>(.*?)</item>", text, re.DOTALL)
    results = []
    for item in items[:lim * 3]:
        title_m = re.search(r"<title>(.*?)</title>", item, re.DOTALL)
        link_m = re.search(r"<link>(.*?)</link>", item, re.DOTALL)
        pub_m = re.search(r"<pubDate>(.*?)</pubDate>", item, re.DOTALL)
        title = (title_m.group(1) if title_m else "").strip()
        link = (link_m.group(1) if link_m else "").strip()
        pub = (pub_m.group(1) if pub_m else "").strip()
        if not title:
            continue
        if ticker and ticker.upper() not in title.upper() and ticker.upper() not in link.upper():
            continue
        results.append({"title": title, "link": link, "pubDate": pub})
        if len(results) >= lim:
            break

    return {"ok": True, "ticker": ticker, "source": source, "items": results, "count": len(results)}


# ---------------------------------------------------------------------------
# listed.coverage.tick
# ---------------------------------------------------------------------------

async def tool_coverage_tick() -> dict:
    conn = await _db()
    try:
        sec_count = await conn.fetchval("SELECT COUNT(*) FROM vertex_isin_security") or 0
        filing_count = await conn.fetchval("SELECT COUNT(*) FROM vertex_isin_filing") or 0
        lei_match_count = await conn.fetchval("SELECT COUNT(*) FROM edge_isin_lei_match") or 0
        ir_doc_count = await conn.fetchval("SELECT COUNT(*) FROM vertex_isin_ir_doc") or 0
        country_rows = await conn.fetch(
            "SELECT country_code, cnt FROM mv_isin_security_country_cnt "
            "ORDER BY cnt DESC"
        )
    finally:
        await conn.close()

    by_country = {r["country_code"]: int(r["cnt"]) for r in country_rows if r["country_code"]}

    return {
        "ok": True,
        "securities": int(sec_count),
        "filings": int(filing_count),
        "leiMatches": int(lei_match_count),
        "irDocs": int(ir_doc_count),
        "byCountry": by_country,
        "snapshotAt": _utc_now(),
    }


# ---------------------------------------------------------------------------
# listed.ingest.hkSecurities
# ---------------------------------------------------------------------------

async def tool_ingest_hk_securities(fromTicker: int = 1, count: int = 25) -> dict:
    """OpenFIGI HKEX main-board tickers 0001-3999 (4-digit zero-padded)."""
    start = max(1, min(int(fromTicker or 1), 3999))
    cnt = max(1, min(int(count or 25), 50))
    tickers = [f"{t:04d}" for t in range(start, min(start + cnt, 4000))]
    if not tickers:
        return {"ok": True, "registered": 0, "exhausted": True}

    FIGI_BATCH = 10
    now = _utc_now()
    registered = skipped = errors = 0

    conn = await _db()
    try:
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(tickers), FIGI_BATCH):
                chunk = tickers[i: i + FIGI_BATCH]
                payload = [
                    {"idType": "TICKER", "idValue": t, "exchCode": "HK", "marketSecDes": "Equity"}
                    for t in chunk
                ]
                try:
                    async with session.post(
                        _OPENFIGI_URL,
                        json=payload,
                        headers={"Content-Type": "application/json"},
                        timeout=aiohttp.ClientTimeout(total=20),
                    ) as resp:
                        if resp.status != 200:
                            continue
                        results = await resp.json()
                except Exception:
                    continue

                for j, entry in enumerate(results):
                    match = (entry.get("data") or [{}])[0]
                    if not match or not match.get("name"):
                        continue
                    ticker = chunk[j]
                    rkey = f"hk-{ticker}"
                    vid = _sec_vid(rkey)
                    try:
                        await conn.execute(
                            "INSERT INTO vertex_isin_security "
                            "(vertex_id, rkey, isin, figi, composite_figi, ticker, name, "
                            "country_code, asset_class, security_type, exch_code, isin_status, "
                            "status, source_did, actor_did, org_did, collected_at, created_at) "
                            "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)",
                            vid, rkey, "",
                            match.get("figi", ""), match.get("compositeFIGI", ""),
                            ticker, match.get("name", ""),
                            "HK", "equity",
                            match.get("securityType", "Common Stock"),
                            "HK", "pending", "active",
                            "did:web:isin.etzhayyim.com:source:openfigi",
                            _OWNER_DID, "anon", now, now,
                        )
                        registered += 1
                    except Exception as e:
                        if _is_dup(e):
                            skipped += 1
                        else:
                            errors += 1

                if i + FIGI_BATCH < len(tickers):
                    await asyncio.sleep(2.6)
    finally:
        await conn.close()

    next_from = (start + cnt) if (start + cnt) <= 3999 else 1
    return {
        "ok": True, "registered": registered, "skipped": skipped, "errors": errors,
        "fromTicker": start, "count": cnt,
        "nextFrom": next_from, "exhausted": (start + cnt) > 3999,
    }


# ---------------------------------------------------------------------------
# listed.ingest.auSecurities
# ---------------------------------------------------------------------------

async def tool_ingest_au_securities(offset: int = 0, limit: int = 200) -> dict:
    """Fetch ASX listed companies CSV then enrich via OpenFIGI."""
    off = max(0, int(offset or 0))
    lim = max(1, min(int(limit or 200), 500))

    async with aiohttp.ClientSession() as session:
        async with session.get(
            _ASX_LIST_URL,
            headers={"User-Agent": _EDGAR_UA},
            timeout=aiohttp.ClientTimeout(total=30),
        ) as resp:
            if resp.status != 200:
                return {"error": f"ASX list {resp.status}", "registered": 0}
            text = await resp.text(errors="replace")

    import csv, io as _io
    rows = []
    reader = csv.reader(_io.StringIO(text))
    for row in reader:
        if len(row) >= 2 and row[1].strip() and row[1].strip().isalpha() and row[0].strip():
            rows.append({"name": row[0].strip(), "ticker": row[1].strip().upper()})

    batch = rows[off: off + lim]
    if not batch:
        return {"ok": True, "registered": 0, "total": len(rows), "exhausted": True}

    FIGI_BATCH = 100
    figi_map: dict[str, dict] = {}
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(batch), FIGI_BATCH):
            chunk = [
                {"idType": "TICKER", "idValue": r["ticker"], "exchCode": "AU", "marketSecDes": "Equity"}
                for r in batch[i: i + FIGI_BATCH]
            ]
            try:
                async with session.post(
                    _OPENFIGI_URL,
                    json=chunk,
                    headers={"Content-Type": "application/json"},
                    timeout=aiohttp.ClientTimeout(total=20),
                ) as r:
                    if r.status == 200:
                        results = await r.json()
                        for j, entry in enumerate(results):
                            match = (entry.get("data") or [{}])[0]
                            if match and (i + j) < len(batch):
                                figi_map[batch[i + j]["ticker"]] = match
            except Exception:
                pass
            if i + FIGI_BATCH < len(batch):
                await asyncio.sleep(2.6)

    now = _utc_now()
    registered = skipped = errors = 0
    conn = await _db()
    try:
        for row in batch:
            ticker = row["ticker"]
            figi = figi_map.get(ticker, {})
            rkey = f"au-{ticker}"
            vid = _sec_vid(rkey)
            try:
                await conn.execute(
                    "INSERT INTO vertex_isin_security "
                    "(vertex_id, rkey, isin, figi, composite_figi, ticker, name, "
                    "country_code, asset_class, security_type, exch_code, isin_status, "
                    "status, source_did, actor_did, org_did, collected_at, created_at) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)",
                    vid, rkey, "",
                    figi.get("figi", ""), figi.get("compositeFIGI", ""),
                    ticker, figi.get("name") or row["name"],
                    "AU", "equity",
                    figi.get("securityType", "Common Stock"),
                    "AU", "pending", "active",
                    "did:web:isin.etzhayyim.com:source:asx",
                    _OWNER_DID, "anon", now, now,
                )
                registered += 1
            except Exception as e:
                if _is_dup(e):
                    skipped += 1
                else:
                    errors += 1
    finally:
        await conn.close()

    return {
        "ok": True, "registered": registered, "skipped": skipped, "errors": errors,
        "total": len(rows), "offset": off, "limit": lim,
        "nextOffset": off + len(batch), "exhausted": len(batch) < lim,
    }


# ---------------------------------------------------------------------------
# listed.ingest.euSecurities
# ---------------------------------------------------------------------------

async def tool_ingest_eu_securities(
    exchCode: str = "GY",
    startCursor: str = "",
    limit: int = 200,
) -> dict:
    """OpenFIGI /v3/filter by exchCode (GY/FP/NA/BB/IM/LN). Paginated via 'next' cursor."""
    exch = str(exchCode or "GY").strip().upper()
    lim = max(1, min(int(limit or 200), 500))
    country = _EXCH_TO_COUNTRY.get(exch, "EU")

    registered = skipped = errors = fetched = 0
    now = _utc_now()
    next_cursor = str(startCursor or "")
    conn = await _db()
    try:
        async with aiohttp.ClientSession() as session:
            while fetched < lim:
                payload: dict[str, Any] = {
                    "exchCode": exch,
                    "marketSector": "Equity",
                }
                if next_cursor:
                    payload["start"] = next_cursor

                try:
                    async with session.post(
                        _OPENFIGI_FILTER_URL,
                        json=payload,
                        headers={"Content-Type": "application/json"},
                        timeout=aiohttp.ClientTimeout(total=20),
                    ) as resp:
                        if resp.status == 429:
                            await asyncio.sleep(5.0)
                            continue
                        if resp.status != 200:
                            break
                        data = await resp.json()
                except Exception:
                    break

                results = data.get("data") or []
                next_cursor = data.get("next") or ""

                for item in results:
                    ticker = item.get("ticker") or ""
                    if not ticker:
                        continue
                    rkey = f"{exch.lower()}-{ticker}"
                    vid = _sec_vid(rkey)
                    try:
                        await conn.execute(
                            "INSERT INTO vertex_isin_security "
                            "(vertex_id, rkey, isin, figi, composite_figi, ticker, name, "
                            "country_code, asset_class, security_type, exch_code, isin_status, "
                            "status, source_did, actor_did, org_did, collected_at, created_at) "
                            "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)",
                            vid, rkey,
                            item.get("isin") or "",
                            item.get("figi", ""), item.get("compositeFigi", ""),
                            ticker, item.get("name", ""),
                            country, "equity",
                            item.get("securityType", "Common Stock"),
                            exch, "pending", "active",
                            f"did:web:isin.etzhayyim.com:source:openfigi:{exch.lower()}",
                            _OWNER_DID, "anon", now, now,
                        )
                        registered += 1
                    except Exception as e:
                        if _is_dup(e):
                            skipped += 1
                        else:
                            errors += 1

                fetched += len(results)
                if not next_cursor or not results:
                    break
                await asyncio.sleep(2.5)
    finally:
        await conn.close()

    return {
        "ok": True, "exchCode": exch, "country": country,
        "registered": registered, "skipped": skipped, "errors": errors,
        "fetched": fetched, "nextCursor": next_cursor,
        "exhausted": not next_cursor,
    }


# ---------------------------------------------------------------------------
# IR document ingest helpers
# ---------------------------------------------------------------------------

_IR_MAX_CHARS = 65536

def _ir_doc_rkey(doc_type: str, url: str) -> str:
    h = hashlib.sha256(url.encode()).hexdigest()[:16]
    return f"{doc_type}-{h}"


def _ir_doc_vid(rkey: str) -> str:
    return f"at://{_OWNER_DID}/com.etzhayyim.apps.isin.irDoc/{rkey}"


async def _save_ir_doc(
    conn: asyncpg.Connection,
    ticker: str,
    company_name: str,
    doc_type: str,
    url: str,
    title: str,
    content_text: str,
    file_size: int,
    page_count: int,
) -> tuple[bool, str | None]:
    rkey = _ir_doc_rkey(doc_type, url)
    vid = _ir_doc_vid(rkey)
    now = _utc_now()
    if content_text and len(content_text) > _IR_MAX_CHARS:
        content_text = content_text[:_IR_MAX_CHARS] + "\n[truncated]"
    try:
        await conn.execute(
            "INSERT INTO vertex_isin_ir_doc "
            "(vertex_id, rkey, ticker, company_name, doc_type, source_url, title, "
            "content_text, file_size_bytes, page_count, actor_did, org_did, created_at) "
            "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)",
            vid, rkey,
            ticker or "", company_name or "",
            doc_type, url, title or "",
            content_text or "", file_size or 0, page_count or 0,
            _OWNER_DID, "anon", now,
        )
        return True, None
    except Exception as e:
        if _is_dup(e):
            return False, "dup"
        return False, str(e)


# ---------------------------------------------------------------------------
# listed.ingest.irPage
# ---------------------------------------------------------------------------

async def tool_ingest_ir_page(
    url: str,
    ticker: str = "",
    companyName: str = "",
) -> dict:
    url = str(url or "").strip()
    if not url:
        return {"error": "url required"}
    ticker = str(ticker or "").strip()
    company_name = str(companyName or "").strip()

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                headers={"User-Agent": _EDGAR_UA},
                timeout=aiohttp.ClientTimeout(total=30),
            ) as resp:
                if resp.status != 200:
                    return {"error": f"fetch {resp.status}", "url": url}
                html = await resp.text(errors="replace")
    except Exception as e:
        return {"error": str(e), "url": url}

    def _parse_html(html_text: str) -> tuple[str, str]:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_text, "lxml")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        t = soup.find("title") or soup.find("h1")
        title_text = t.get_text(strip=True) if t else ""
        body = soup.get_text(separator="\n", strip=True)
        return title_text, body

    title, content = await asyncio.to_thread(_parse_html, html)

    conn = await _db()
    try:
        saved, err = await _save_ir_doc(
            conn, ticker, company_name, "ir_page", url,
            title, content, len(html.encode()), 0,
        )
    finally:
        await conn.close()

    return {
        "ok": True, "url": url, "ticker": ticker,
        "title": title, "chars": len(content),
        "saved": saved, "note": err,
    }


# ---------------------------------------------------------------------------
# listed.ingest.irPdf
# ---------------------------------------------------------------------------

async def tool_ingest_ir_pdf(
    url: str,
    ticker: str = "",
    companyName: str = "",
) -> dict:
    url = str(url or "").strip()
    if not url:
        return {"error": "url required"}
    ticker = str(ticker or "").strip()
    company_name = str(companyName or "").strip()

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                headers={"User-Agent": _EDGAR_UA},
                timeout=aiohttp.ClientTimeout(total=60),
            ) as resp:
                if resp.status != 200:
                    return {"error": f"fetch {resp.status}", "url": url}
                data = await resp.read()
    except Exception as e:
        return {"error": str(e), "url": url}

    def _parse_pdf(pdf_bytes: bytes) -> tuple[str, int]:
        import pdfplumber
        texts: list[str] = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages[:50]:
                t = page.extract_text() or ""
                if t:
                    texts.append(t)
        return "\n".join(texts), page_count

    try:
        content, page_count = await asyncio.to_thread(_parse_pdf, data)
    except Exception as e:
        return {"error": f"PDF parse: {e}", "url": url}

    title = url.rsplit("/", 1)[-1].replace(".pdf", "").replace("_", " ")

    conn = await _db()
    try:
        saved, err = await _save_ir_doc(
            conn, ticker, company_name, "ir_pdf", url,
            title, content, len(data), page_count,
        )
    finally:
        await conn.close()

    return {
        "ok": True, "url": url, "ticker": ticker,
        "pages": page_count, "chars": len(content),
        "saved": saved, "note": err,
    }


# ---------------------------------------------------------------------------
# listed.ingest.irExcel
# ---------------------------------------------------------------------------

async def tool_ingest_ir_excel(
    url: str,
    ticker: str = "",
    companyName: str = "",
) -> dict:
    url = str(url or "").strip()
    if not url:
        return {"error": "url required"}
    ticker = str(ticker or "").strip()
    company_name = str(companyName or "").strip()

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                headers={"User-Agent": _EDGAR_UA},
                timeout=aiohttp.ClientTimeout(total=60),
            ) as resp:
                if resp.status != 200:
                    return {"error": f"fetch {resp.status}", "url": url}
                data = await resp.read()
    except Exception as e:
        return {"error": str(e), "url": url}

    def _parse_excel(xlsx_bytes: bytes) -> tuple[str, int]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet_name in wb.sheetnames[:10]:
            ws = wb[sheet_name]
            lines.append(f"=== {sheet_name} ===")
            for row in ws.iter_rows(max_row=200, values_only=True):
                row_vals = [str(c) if c is not None else "" for c in row]
                line = "\t".join(row_vals).strip()
                if line:
                    lines.append(line)
        return "\n".join(lines), len(wb.sheetnames)

    try:
        content, sheet_count = await asyncio.to_thread(_parse_excel, data)
    except Exception as e:
        return {"error": f"Excel parse: {e}", "url": url}

    title = url.rsplit("/", 1)[-1].replace(".xlsx", "").replace("_", " ")

    conn = await _db()
    try:
        saved, err = await _save_ir_doc(
            conn, ticker, company_name, "ir_excel", url,
            title, content, len(data), sheet_count,
        )
    finally:
        await conn.close()

    return {
        "ok": True, "url": url, "ticker": ticker,
        "sheets": sheet_count, "chars": len(content),
        "saved": saved, "note": err,
    }


# ---------------------------------------------------------------------------
# listed.ingest.irWord
# ---------------------------------------------------------------------------

async def tool_ingest_ir_word(
    url: str,
    ticker: str = "",
    companyName: str = "",
) -> dict:
    url = str(url or "").strip()
    if not url:
        return {"error": "url required"}
    ticker = str(ticker or "").strip()
    company_name = str(companyName or "").strip()

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                headers={"User-Agent": _EDGAR_UA},
                timeout=aiohttp.ClientTimeout(total=60),
            ) as resp:
                if resp.status != 200:
                    return {"error": f"fetch {resp.status}", "url": url}
                data = await resp.read()
    except Exception as e:
        return {"error": str(e), "url": url}

    def _parse_word(docx_bytes: bytes) -> tuple[str, int]:
        import docx
        doc = docx.Document(io.BytesIO(docx_bytes))
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paras), len(paras)

    try:
        content, para_count = await asyncio.to_thread(_parse_word, data)
    except Exception as e:
        return {"error": f"Word parse: {e}", "url": url}

    title = url.rsplit("/", 1)[-1].replace(".docx", "").replace("_", " ")

    conn = await _db()
    try:
        saved, err = await _save_ir_doc(
            conn, ticker, company_name, "ir_word", url,
            title, content, len(data), para_count,
        )
    finally:
        await conn.close()

    return {
        "ok": True, "url": url, "ticker": ticker,
        "paragraphs": para_count, "chars": len(content),
        "saved": saved, "note": err,
    }


# ---------------------------------------------------------------------------
# FastAPI LangServer
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# listed.normalize.linkSecurityFiling
# ---------------------------------------------------------------------------

async def tool_link_security_filing(limit: int = 200) -> dict:
    """Insert edge_isin_security_filing by matching ticker across both tables."""
    lim = max(1, min(int(limit or 200), 500))
    conn = await _db()
    try:
        filings = await conn.fetch(
            f"SELECT vertex_id, ticker FROM vertex_isin_filing "
            f"WHERE ticker IS NOT NULL AND ticker != '' "
            f"ORDER BY created_at DESC LIMIT {lim}"
        )
        linked = skipped = 0
        now = _utc_now()
        for fil in filings:
            sec_row = await conn.fetchrow(
                "SELECT vertex_id FROM vertex_isin_security "
                "WHERE ticker = $1 LIMIT 1",
                fil["ticker"],
            )
            if not sec_row:
                continue
            edge_id = (
                f"at://{_OWNER_DID}/com.etzhayyim.apps.isin.securityFiling"
                f"/{fil['ticker']}-{uuid.uuid4().hex[:8]}"
            )
            try:
                await conn.execute(
                    "INSERT INTO edge_isin_security_filing "
                    "(edge_id, src_vid, dst_vid, ticker, actor_did, org_did, created_at) "
                    "VALUES ($1,$2,$3,$4,$5,$6,$7)",
                    edge_id, sec_row["vertex_id"], fil["vertex_id"],
                    fil["ticker"], _OWNER_DID, "anon", now,
                )
                linked += 1
            except Exception as e:
                if _is_dup(e):
                    skipped += 1
    finally:
        await conn.close()
    return {"ok": True, "linked": linked, "skipped": skipped}


# ---------------------------------------------------------------------------
# Rolling sweep cursors (in-memory; resets to 0 on pod restart, which is fine
# because all inserts are dedup-safe via UniqueViolationError).
# ---------------------------------------------------------------------------

_us_cursor: int = 0
_jp_cursor: int = 1000
_hk_cursor: int = 1
_au_cursor: int = 0
_eu_exch_idx: int = 0
_eu_cursors: dict[str, str] = {}  # exchCode → next_cursor
_edinet_cursor: int = 0


# ---------------------------------------------------------------------------
# Cron batch runners — called by k8s CronJob via POST /cron/*
# Each returns immediately with {"ok": true, "started": true} after launching
# an asyncio background task so the curl trigger never times out.
# ---------------------------------------------------------------------------

async def _run_us_sweep(batch_size: int) -> None:
    global _us_cursor
    result = await tool_ingest_us_securities(offset=_us_cursor, limit=batch_size, enrichFigi=True)
    total = result.get("total") or 0
    _us_cursor = ((_us_cursor + batch_size) % total) if total else 0


async def _run_jp_sweep(batch_size: int) -> None:
    global _jp_cursor
    result = await tool_ingest_jp_securities(fromTicker=_jp_cursor, count=batch_size)
    if result.get("exhausted"):
        _jp_cursor = 1000
    else:
        _jp_cursor = result.get("nextFrom") or (_jp_cursor + batch_size)


async def _run_enrich_cik(limit: int) -> None:
    conn = await _db()
    try:
        rows = await conn.fetch(
            f"SELECT cik FROM vertex_isin_security "
            f"WHERE cik IS NOT NULL AND cik != '' "
            f"AND (exchange_mic IS NULL OR exchange_mic = '') "
            f"AND country_code = 'US' "
            f"ORDER BY created_at DESC LIMIT {limit}"
        )
    finally:
        await conn.close()
    for row in rows:
        cik_val = row["cik"]
        if cik_val and str(cik_val).isdigit():
            await tool_enrich_cik(cik=int(cik_val))
            await asyncio.sleep(0.15)


async def _run_link_lei(limit: int) -> None:
    await tool_link_lei_isin(limit=limit)


async def _run_link_filing(limit: int) -> None:
    await tool_link_security_filing(limit=limit)


async def _run_hk_sweep(batch_size: int) -> None:
    global _hk_cursor
    result = await tool_ingest_hk_securities(fromTicker=_hk_cursor, count=batch_size)
    if result.get("exhausted"):
        _hk_cursor = 1
    else:
        _hk_cursor = result.get("nextFrom") or (_hk_cursor + batch_size)


async def _run_au_sweep(batch_size: int) -> None:
    global _au_cursor
    result = await tool_ingest_au_securities(offset=_au_cursor, limit=batch_size)
    total = result.get("total") or 0
    if result.get("exhausted") or not total:
        _au_cursor = 0
    else:
        _au_cursor = ((_au_cursor + batch_size) % total)


async def _run_edinet_sweep(limit: int) -> None:
    await tool_ingest_edinet_sweep(limit=limit)


async def _run_eu_sweep(batch_size: int) -> None:
    global _eu_exch_idx, _eu_cursors
    exch = _EU_EXCH_CODES[_eu_exch_idx % len(_EU_EXCH_CODES)]
    start_cursor = _eu_cursors.get(exch, "")
    result = await tool_ingest_eu_securities(
        exchCode=exch, startCursor=start_cursor, limit=batch_size,
    )
    if result.get("exhausted"):
        _eu_cursors.pop(exch, None)
        _eu_exch_idx += 1
    else:
        _eu_cursors[exch] = result.get("nextCursor") or ""


async def _run_bootstrap() -> None:
    """Full US+JP+HK+AU+EU sweep. Sequential to avoid OpenFIGI rate-limit double-hits."""
    global _us_cursor, _jp_cursor, _hk_cursor, _au_cursor, _eu_exch_idx, _eu_cursors

    try:
        first = await tool_ingest_us_securities(offset=0, limit=500, enrichFigi=True)
        total = first.get("total") or 0
        for off in range(500, total, 500):
            await tool_ingest_us_securities(offset=off, limit=500, enrichFigi=True)
            await asyncio.sleep(1.0)
        _us_cursor = 0
    except Exception:
        pass

    try:
        from_t = 1000
        while from_t < 10000:
            result = await tool_ingest_jp_securities(fromTicker=from_t, count=50)
            if result.get("exhausted"):
                _jp_cursor = 1000
                break
            from_t = result.get("nextFrom") or (from_t + 50)
            _jp_cursor = from_t
            await asyncio.sleep(0.5)
    except Exception:
        pass

    # HK: 0001-3999
    try:
        hk_t = 1
        while hk_t < 4000:
            result = await tool_ingest_hk_securities(fromTicker=hk_t, count=50)
            if result.get("exhausted"):
                _hk_cursor = 1
                break
            hk_t = result.get("nextFrom") or (hk_t + 50)
            _hk_cursor = hk_t
            await asyncio.sleep(0.5)
    except Exception:
        pass

    # AU: full ASX list
    try:
        au_off = 0
        while True:
            result = await tool_ingest_au_securities(offset=au_off, limit=200)
            if result.get("exhausted"):
                _au_cursor = 0
                break
            au_off = result.get("nextOffset") or (au_off + 200)
            _au_cursor = au_off
            await asyncio.sleep(1.0)
    except Exception:
        pass

    # EU: all exchCodes
    try:
        for exch in _EU_EXCH_CODES:
            cursor = ""
            while True:
                result = await tool_ingest_eu_securities(
                    exchCode=exch, startCursor=cursor, limit=200,
                )
                if result.get("exhausted") or not result.get("nextCursor"):
                    _eu_cursors.pop(exch, None)
                    break
                cursor = result.get("nextCursor") or ""
                _eu_cursors[exch] = cursor
                await asyncio.sleep(2.5)
        _eu_exch_idx = 0
    except Exception:
        pass


TOOLS: dict[str, Any] = {
    "listed.ingest.usSecurities": tool_ingest_us_securities,
    "listed.ingest.jpSecurities": tool_ingest_jp_securities,
    "listed.ingest.hkSecurities": tool_ingest_hk_securities,
    "listed.ingest.auSecurities": tool_ingest_au_securities,
    "listed.ingest.euSecurities": tool_ingest_eu_securities,
    "listed.enrich.cik": tool_enrich_cik,
    "listed.ingest.edinetFiling": tool_ingest_edinet_filing,
    "listed.ingest.edinetPdf": tool_ingest_edinet_pdf,
    "listed.ingest.edinetSweep": tool_ingest_edinet_sweep,
    "listed.normalize.linkLeiIsin": tool_link_lei_isin,
    "listed.normalize.linkSecurityFiling": tool_link_security_filing,
    "listed.news.fetchPressReleases": tool_fetch_press_releases,
    "listed.ingest.irPage": tool_ingest_ir_page,
    "listed.ingest.irPdf": tool_ingest_ir_pdf,
    "listed.ingest.irExcel": tool_ingest_ir_excel,
    "listed.ingest.irWord": tool_ingest_ir_word,
    "listed.coverage.tick": tool_coverage_tick,
}

app = FastAPI(title="lg-isin", version="1.0.0")


@app.get("/healthz")
async def healthz() -> dict[str, Any]:
    return {
        "ok": True,
        "service": "lg-isin",
        "runtimeKind": "k8s-langserver",
        "tools": sorted(TOOLS),
        "edinetKeyConfigured": bool(EDINET_KEY),
        "sweepCursors": {
            "us": _us_cursor, "jp": _jp_cursor,
            "hk": _hk_cursor, "au": _au_cursor,
            "eu_exch": _EU_EXCH_CODES[_eu_exch_idx % len(_EU_EXCH_CODES)],
            "edinet": _edinet_cursor,
        },
    }


@app.get("/health")
async def health() -> dict[str, Any]:
    return {"ok": True}


@app.get("/tools")
async def tools_list() -> dict[str, Any]:
    return {"tools": [{"name": n, "runtime": "langserver"} for n in sorted(TOOLS)]}


async def _invoke(name: str, arguments: dict[str, Any]) -> Any:
    fn = TOOLS.get(name)
    if fn is None:
        raise HTTPException(status_code=404, detail=f"unknown tool: {name}")
    return await fn(**arguments)


@app.post("/invoke")
async def invoke(payload: dict[str, Any]) -> dict[str, Any]:
    name = str(payload.get("name") or payload.get("tool") or "")
    arguments = payload.get("arguments") or payload.get("input") or {}
    if not isinstance(arguments, dict):
        raise HTTPException(status_code=400, detail="arguments must be an object")
    return {"ok": True, "name": name, "result": await _invoke(name, arguments)}


@app.post("/runs")
async def runs(payload: dict[str, Any]) -> dict[str, Any]:
    assistant_id = str(payload.get("assistant_id") or "")
    arguments = payload.get("input") or payload.get("arguments") or {}
    if not isinstance(arguments, dict):
        raise HTTPException(status_code=400, detail="input must be an object")
    return {
        "status": "completed",
        "assistant_id": assistant_id,
        "output": await _invoke(assistant_id, arguments),
    }


@app.post("/cron/sweep-us")
async def cron_sweep_us(payload: dict[str, Any] = {}) -> dict[str, Any]:
    batch = max(50, min(int((payload or {}).get("batchSize") or 500), 1000))
    asyncio.create_task(_run_us_sweep(batch))
    return {"ok": True, "started": True, "cursor": _us_cursor, "batchSize": batch}


@app.post("/cron/sweep-jp")
async def cron_sweep_jp(payload: dict[str, Any] = {}) -> dict[str, Any]:
    batch = max(10, min(int((payload or {}).get("batchSize") or 50), 200))
    asyncio.create_task(_run_jp_sweep(batch))
    return {"ok": True, "started": True, "cursor": _jp_cursor, "batchSize": batch}


@app.post("/cron/enrich-cik")
async def cron_enrich_cik(payload: dict[str, Any] = {}) -> dict[str, Any]:
    limit = max(1, min(int((payload or {}).get("limit") or 20), 100))
    asyncio.create_task(_run_enrich_cik(limit))
    return {"ok": True, "started": True, "limit": limit}


@app.post("/cron/link-lei")
async def cron_link_lei(payload: dict[str, Any] = {}) -> dict[str, Any]:
    limit = max(1, min(int((payload or {}).get("limit") or 200), 500))
    asyncio.create_task(_run_link_lei(limit))
    return {"ok": True, "started": True, "limit": limit}


@app.post("/cron/sweep-hk")
async def cron_sweep_hk(payload: dict[str, Any] = {}) -> dict[str, Any]:
    batch = max(10, min(int((payload or {}).get("batchSize") or 50), 200))
    asyncio.create_task(_run_hk_sweep(batch))
    return {"ok": True, "started": True, "cursor": _hk_cursor, "batchSize": batch}


@app.post("/cron/sweep-au")
async def cron_sweep_au(payload: dict[str, Any] = {}) -> dict[str, Any]:
    batch = max(50, min(int((payload or {}).get("batchSize") or 200), 500))
    asyncio.create_task(_run_au_sweep(batch))
    return {"ok": True, "started": True, "cursor": _au_cursor, "batchSize": batch}


@app.post("/cron/sweep-eu")
async def cron_sweep_eu(payload: dict[str, Any] = {}) -> dict[str, Any]:
    batch = max(50, min(int((payload or {}).get("batchSize") or 200), 500))
    asyncio.create_task(_run_eu_sweep(batch))
    exch = _EU_EXCH_CODES[_eu_exch_idx % len(_EU_EXCH_CODES)]
    return {"ok": True, "started": True, "exchCode": exch, "batchSize": batch}


@app.post("/cron/edinet-sweep")
async def cron_edinet_sweep(payload: dict[str, Any] = {}) -> dict[str, Any]:
    limit = max(1, min(int((payload or {}).get("limit") or 10), 50))
    asyncio.create_task(_run_edinet_sweep(limit))
    return {
        "ok": True, "started": True, "limit": limit,
        "cursor": _edinet_cursor, "edinetKeyConfigured": bool(EDINET_KEY),
    }


@app.post("/cron/link-filing")
async def cron_link_filing(payload: dict[str, Any] = {}) -> dict[str, Any]:
    limit = max(1, min(int((payload or {}).get("limit") or 200), 500))
    asyncio.create_task(_run_link_filing(limit))
    return {"ok": True, "started": True, "limit": limit}


@app.post("/cron/bootstrap")
async def cron_bootstrap() -> dict[str, Any]:
    asyncio.create_task(_run_bootstrap())
    return {
        "ok": True, "started": True,
        "note": "US full sweep (~4 min) then JP full sweep (~40 min) running in background",
    }


@app.post("/cron/coverage")
async def cron_coverage() -> dict[str, Any]:
    result = await tool_coverage_tick()
    return result


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8080"))
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")
